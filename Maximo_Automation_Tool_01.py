import streamlit as st
import pandas as pd
import threading
import time
import os
import io
import sys
import json
import shutil
import datetime
import logging
import random
from pathlib import Path
from typing import Optional
from cryptography.fernet import Fernet
import asyncio
import subprocess

# Ensure Windows event loop supports subprocess in threads (needed by Playwright)
if sys.platform.startswith("win"):
    try:
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    except Exception:
        pass

# Optional Playwright imports (guarded)
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
    PLAYWRIGHT_AVAILABLE = True
except Exception:
    PLAYWRIGHT_AVAILABLE = False

# -----------------------
# Page setup and styling
# -----------------------
st.set_page_config(
    page_title="Maximo Automation Tool - Improved",
    page_icon="🧩",
    layout="centered",
    initial_sidebar_state="expanded"
)

CUSTOM_CSS = """
<style>
/* Base colors and font */
:root {
  --bg: #ffffff;
  --text: #1a1a1a;
  --subtle: #404040;
  --accent: #1a1a1a;
}
html, body, [data-testid="stAppViewContainer"] {
  background: var(--bg) !important;
  font-family: "Segoe UI", system-ui, -apple-system, "Helvetica Neue", Arial, sans-serif;
  color: var(--text);
}
h1, h2, h3 { color: var(--text); }
.block-container { padding-top: 1rem; padding-bottom: 2rem; }

/* Cards */
.card {
    background: #f2f2f2; /* CTk light vibe */
    border: 2px solid #d9d9d9;
    border-radius: 15px; /* match CTk rounding */
  padding: 16px 18px;
  margin-bottom: 16px;
}
.card-title {
  font-weight: 700;
  font-size: 18px;
  margin-bottom: 8px;
}

/* Buttons wrap */
.button-row {
  display: flex;
  gap: 8px;
  flex-wrap: wrap;
}

/* CTk-like button styling (applies to all Streamlit buttons) */
.stButton > button {
    border-radius: 8px;
    font-weight: 700;
    border: 1px solid #1a1a1a10;
    transition: background 0.15s ease, transform 0.02s ease;
}
.stButton > button:hover { transform: translateY(-1px); }
/* Primary style mimic for Start */
.stButton > button[kind="primary"], .stButton > button:has(span:contains("Start")) {
    background: #007bff !important;
    color: #fff !important;
}
.stButton > button[kind="primary"]:hover, .stButton > button:has(span:contains("Start")):hover { background: #0056b3 !important; }
/* Neutral style for Pause/Save Logs */
.stButton > button:has(span:contains("Pause")),
.stButton > button:has(span:contains("Save Logs")) {
    background: #6c757d !important;
    color: #fff !important;
}
.stButton > button:has(span:contains("Pause")):hover,
.stButton > button:has(span:contains("Save Logs")):hover { background: #5a6268 !important; }
/* Danger style for Stop */
.stButton > button:has(span:contains("Stop")) {
    background: #dc3545 !important;
    color: #fff !important;
}
.stButton > button:has(span:contains("Stop")):hover { background: #bd2130 !important; }

/* Log area */
.log-container {
  background: #111;
  color: #d4d4d4;
  border-radius: 8px;
  border: 1px solid #333;
    padding: 8px;
    max-height: 280px;
    overflow-y: auto;
}
.small-muted { color: #808080; font-size: 13px; }

/* Colored log lines */
.log-line { display: block; font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; white-space: pre-wrap; }
.log-info { color: #4CAF50; }
.log-warning { color: #FFA500; }
.log-error { color: #FF4D4F; }

/* Inputs alignment */
label, .stTextInput label, .stCheckbox label { color: var(--text) !important; }

/* Progress text */
.progress-text { font-weight: 600; }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# -----------------------
# Constants and files
# -----------------------
APP_TITLE = "Maximo Automation Tool - Improved"
APP_VERSION = "1.00"
KEY_FILENAME = ".maximo_key"
CONFIG_FILENAME = ".maximo_config.json"
LOG_DIRNAME = "maximo_logs"

DEV_URL = "https://system.arsal.com/maxarsaldev/webclient/login/login.jsp"
PROD_URL = "https://system.arsal.com/maxarsal/webclient/login/login.jsp"

HOME = Path.home()
CONFIG_FILE = HOME / CONFIG_FILENAME
KEY_FILE = HOME / KEY_FILENAME
LOG_DIR = HOME / LOG_DIRNAME
LOG_DIR.mkdir(parents=True, exist_ok=True)

# -----------------------
# Logger (to file + memory)
# -----------------------
class InMemoryLogHandler(logging.Handler):
    def emit(self, record):
        try:
            msg = self.format(record)
            logs = st.session_state.get("logs", [])
            logs.append(msg)
            st.session_state["logs"] = logs
        except Exception:
            # When logging from background threads without Streamlit context
            # just drop in-memory UI logs to avoid ScriptRunContext warnings
            pass

def setup_logger():
    logger = logging.getLogger("MaximoAutomation")
    logger.setLevel(logging.DEBUG)
    logger.handlers = []

    # File handler (daily)
    today = datetime.datetime.now().strftime("%Y%m%d")
    log_file = LOG_DIR / f"maximo_automation_{today}.log"
    fh = logging.FileHandler(log_file, encoding="utf-8", mode="a")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "[%(asctime)s] - %(levelname)-8s - %(threadName)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"))
    logger.addHandler(fh)

    # In-memory for UI
    ih = InMemoryLogHandler()
    ih.setLevel(logging.INFO)
    ih.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s", datefmt="%H:%M:%S"))
    logger.addHandler(ih)

    logger.info(f"Starting {APP_TITLE} v{APP_VERSION}")
    logger.info(f"Log file: {log_file}")
    try:
        logger.info(f"Playwright available: {PLAYWRIGHT_AVAILABLE}")
    except Exception:
        pass
    return logger

# -----------------------
# Credentials encryption
# -----------------------
def ensure_key():
    if not KEY_FILE.exists():
        key = Fernet.generate_key()
        KEY_FILE.write_bytes(key)
        try:
            os.chmod(KEY_FILE, 0o600)
        except Exception:
            pass
    return KEY_FILE.read_bytes()

def get_cipher():
    key = ensure_key()
    return Fernet(key)

def save_credentials(username: str, password: str, url: str, remember: bool):
    try:
        if not remember:
            if CONFIG_FILE.exists():
                CONFIG_FILE.unlink()
            st.session_state["status"] = "Saved credentials removed."
            return True
        cipher = get_cipher()
        enc_pwd = cipher.encrypt(password.encode("utf-8")).decode("utf-8")
        CONFIG_FILE.write_text(json.dumps({
            "username": username,
            "password": enc_pwd,
            "url": url,
            "remember": True
        }, ensure_ascii=False, indent=2), encoding="utf-8")
        st.session_state["status"] = "Credentials saved."
        return True
    except Exception as e:
        st.session_state["status"] = f"Error saving credentials: {e}"
        return False

def load_credentials():
    if not CONFIG_FILE.exists():
        return "", "", DEV_URL, False
    try:
        cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        cipher = get_cipher()
        dec_pwd = cipher.decrypt(cfg.get("password", "").encode("utf-8")).decode("utf-8") if cfg.get("password") else ""
        return cfg.get("username", ""), dec_pwd, cfg.get("url", DEV_URL), bool(cfg.get("remember", False))
    except Exception:
        return "", "", DEV_URL, False

# -----------------------
# Session state bootstrap
# -----------------------
def init_state():
    defaults = {
        "url": DEV_URL,
        "username": "",
        "password": "",
        "remember": False,
        "show_password": False,
        "show_browser": True,
        "dev_env": True,
        "prod_env": False,
        "excel_file": None,
        "excel_path": None,
        "working_excel_path": None,
        "progress": 0.0,
        "progress_text": "0%",
        "status": "",
        "logs": [],
        "automation_thread": None,
        "stop_event": None,
        "paused": False,
        "is_running": False,
        "total_sheets": 0,
        "current_sheet": None,
        "excel_cache": None
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
    # New defaults for parity enhancements
    if "flow_mode" not in st.session_state:
        st.session_state["flow_mode"] = "add_services"  # or "edit_prices"
    if "last_wo_number" not in st.session_state:
        st.session_state["last_wo_number"] = ""
    if "last_bill_batch" not in st.session_state:
        st.session_state["last_bill_batch"] = ""
    # Step selection defaults (ordered pipeline)
    if "selected_steps" not in st.session_state:
        st.session_state["selected_steps"] = [
            "add_services",
            "route_to_comp",
            "create_bill",
            "put_prices",
        ]
    if "start_at" not in st.session_state:
        st.session_state["start_at"] = "add_services"
    # Always show browser; ensure no dry-run by default
    st.session_state["show_browser"] = True
    if "dry_run" in st.session_state:
        st.session_state["dry_run"] = False
    # Load saved credentials if any
    u, p, url, rem = load_credentials()
    if u and not st.session_state["username"]:
        st.session_state["username"] = u
        st.session_state["password"] = p
        st.session_state["url"] = url
        st.session_state["remember"] = rem

init_state()
logger = setup_logger()

# -----------------------
# UI helpers
# -----------------------
def append_log(msg: str, level="info"):
    if level == "error":
        logger.error(msg)
    elif level == "warning":
        logger.warning(msg)
    elif level == "debug":
        logger.debug(msg)
    else:
        logger.info(msg)

def update_status(msg: str, level="info"):
    st.session_state["status"] = msg
    append_log(msg, level=level)

def set_progress(value: float):
    val = max(0.0, min(1.0, float(value)))
    st.session_state["progress"] = val
    st.session_state["progress_text"] = f"{int(val * 100)}%"

def disable_controls():
    st.session_state["is_running"] = True

def enable_controls():
    st.session_state["is_running"] = False

# -----------------------
# Automation adapter (preserve original function names/logic)
# -----------------------
class VarProxy:
    def __init__(self, key: str, fallback=None):
        self.key = key
        self.fallback = fallback
    def get(self):
        try:
            return st.session_state.get(self.key, self.fallback)
        except Exception:
            # Accessing session_state from a non-Streamlit thread can raise; use fallback
            return self.fallback
    def set(self, value):
        try:
            st.session_state[self.key] = value
        except Exception:
            # If not in Streamlit context, at least keep the latest value locally
            self.fallback = value

class StreamlitAutomation:
    """Adapter class to host original automation methods with minimal changes."""
    def __init__(self,
                 url: Optional[str] = None,
                 username: Optional[str] = None,
                 password: Optional[str] = None,
                 show_browser: Optional[bool] = True,
                 remember: Optional[bool] = False,
                 stop_event: Optional[threading.Event] = None,
                 paused: bool = False,
                 excel_path: Optional[str] = None,
                 working_excel_path: Optional[str] = None,
                 excel_cache=None):
        # Proxies to mimic Tkinter StringVar/BooleanVar with thread-safe fallbacks
        self.url_var = VarProxy("url", fallback=url)
        self.username_var = VarProxy("username", fallback=username)
        self.password_var = VarProxy("password", fallback=password)
        self.show_browser_var = VarProxy("show_browser", fallback=show_browser)
        self.remember_var = VarProxy("remember", fallback=remember)

        # State fields expected by original code (avoid session_state in threads)
        self.logger = logging.getLogger("MaximoAutomation")
        self.stop_event = stop_event or threading.Event()
        self.paused = paused
        self.excel_path = excel_path
        self.working_excel_path = working_excel_path or excel_path
        self._excel_file_cache = excel_cache

        # Placeholders for Playwright
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

        # Data placeholders
        self.current_sheet = None
        self.wo_number = None
        self.wo_description = None
        self.excel_original_description = None
        self.bill_batch = None
        self.bill_status = None
        self.reviewed_by = None
        self.services_data = None

    # Thin wrappers to reuse existing UI helpers
    def append_log(self, msg: str, level: str = "info"):
        append_log(msg, level)
    def update_status(self, msg: str, level: str = "info"):
        update_status(msg, level)

    # Pause helpers
    def check_pause(self):
        if self.paused:
            self.wait_if_paused()
    def wait_if_paused(self):
        while self.paused and not self.stop_event.is_set():
            time.sleep(0.5)

    # --- Copied logic methods (minimal adaptation) ---
    def update_excel_file(self):
        """Write results back to Excel directly to the working copy (preserve macros)."""
        try:
            if not hasattr(self, 'working_excel_path') or not self.working_excel_path:
                self.logger.warning("No Excel working copy path available")
                return False

            from openpyxl import load_workbook
            is_xlsm = str(Path(self.working_excel_path).suffix).lower() == ".xlsm"
            wb = load_workbook(self.working_excel_path, keep_vba=is_xlsm)

            # Get current sheet
            if hasattr(self, 'current_sheet') and self.current_sheet in wb.sheetnames:
                ws = wb[self.current_sheet]
            else:
                ws = wb.active

            # Update header information (B2-B6)
            if hasattr(self, 'wo_number') and self.wo_number:
                ws['B2'] = self.wo_number
            if hasattr(self, 'wo_description') and self.wo_description:
                ws['B3'] = self.wo_description
            if hasattr(self, 'bill_batch') and self.bill_batch:
                ws['B4'] = self.bill_batch
            if hasattr(self, 'bill_status') and self.bill_status:
                ws['B5'] = self.bill_status
            else:
                ws['B5'] = "CREATED"
            if hasattr(self, 'reviewed_by') and self.reviewed_by:
                ws['B6'] = self.reviewed_by

            wb.save(self.working_excel_path)
            wb.close()

            self.append_log(f"Excel updated successfully (Sheet: {ws.title})")
            self.append_log(
                f"Updated: WO={getattr(self, 'wo_number', 'N/A')}, Description={getattr(self, 'wo_description', 'N/A')}, Bill Batch={getattr(self, 'bill_batch', 'N/A')}"
            )
            self.update_status("Excel updated", "info")
            return True
        except Exception as e:
            self.logger.exception("Failed to update excel")
            self.update_status(f"Failed to update Excel: {e}", "error")
            return False

    def enter_services_prices(self, page):
        """Enter new prices in Services tab from Excel data with multi-page support"""
        try:
            self.check_pause()
            self.append_log("Entering new service prices from Excel...")

            services_tab = page.locator('text="Services"').first
            services_tab.click()
            time.sleep(2)

            page_number = 1
            processed_services = set()
            updated_count = 0
            total_services = 0

            excel_services_map = {}
            if hasattr(self, 'services_data') and self.services_data is not None and not self.services_data.empty:
                for _, service in self.services_data.iterrows():
                    service_item = service['Service Item']
                    total_price = service['Total Price']
                    if service_item and total_price is not None:
                        excel_services_map[service_item] = total_price

            while True:
                self.check_pause()
                service_elements = page.evaluate("""
                    () => {
                        const rows = Array.from(document.querySelectorAll('tr[id*="_SERVICES_"]'));
                        return rows.map(row => {
                            const serviceSpan = row.querySelector('span[wrapoff="true"][title^="IS"]');
                            const priceInput = row.querySelector('td[headers*="_tdrow_[C:8]"] input[role="textbox"]');
                            return {
                                serviceCode: serviceSpan ? serviceSpan.getAttribute('title') : null,
                                inputId: priceInput ? priceInput.id : null
                            };
                        }).filter(item => item.serviceCode && item.inputId);
                    }
                """)

                if not service_elements:
                    self.logger.warning(f"No service items found on page {page_number}")
                    break

                total_services += len(service_elements)
                self.logger.info(f"Processing {len(service_elements)} services on page {page_number}")

                for service in service_elements:
                    self.check_pause()
                    try:
                        service_code = service['serviceCode']
                        if not service_code or service_code in processed_services:
                            continue

                        price_input = page.locator(f"#{service['inputId']}").first
                        if not price_input:
                            continue

                        excel_price = excel_services_map.get(service_code)
                        if excel_price is not None:
                            new_value = str(excel_price)
                            success = page.evaluate("""
                                (input, value) => {
                                    try {
                                        input.value = value;
                                        input.dispatchEvent(new Event('change', { bubbles: true }));
                                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                                        Object.getOwnPropertyDescriptor(
                                            window.HTMLInputElement.prototype, 'value'
                                        ).set.call(input, value);
                                        input.dispatchEvent(new Event('input', { bubbles: true }));
                                        input.dispatchEvent(new KeyboardEvent('keyup', { key: 'Tab' }));
                                        return input.value === value;
                                    } catch (e) {
                                        console.error('Error setting value:', e);
                                        return false;
                                    }
                                }
                            """, price_input, new_value)

                            if success:
                                updated_count += 1
                                self.logger.info(f"Set price for {service_code} to {new_value}")
                            else:
                                self.logger.warning(f"Failed to set price for {service_code}")
                        else:
                            page.evaluate("""
                                (input) => {
                                    input.value = '0.00';
                                    input.dispatchEvent(new Event('change', { bubbles: true }));
                                    input.dispatchEvent(new Event('input', { bubbles: true }));
                                    input.dispatchEvent(new Event('blur', { bubbles: true }));
                                }
                            """, price_input)
                            self.logger.info(f"Service {service_code} not in Excel, set to 0.00")
                            updated_count += 1

                        processed_services.add(service_code)
                        time.sleep(0.2)
                    except Exception as e:
                        self.logger.error(f"Error processing service {service_code}: {str(e)}")
                        continue

                next_button = page.locator('img[id$="-ti7_img"]').first
                if not next_button or not next_button.is_visible() or 'off.gif' in (next_button.get_attribute('src') or ''):
                    self.logger.info("No more pages to process")
                    break
                try:
                    self.check_pause()
                    next_button.click()
                    page.wait_for_load_state("networkidle", timeout=30000)
                    time.sleep(1.5)
                    page_number += 1
                except Exception as e:
                    self.logger.error(f"Error navigating to next page: {str(e)}")
                    break

            self.append_log(
                f"Completed price updates:\n" \
                f"- Total pages processed: {page_number}\n" \
                f"- Total services found: {total_services}\n" \
                f"- Successfully updated: {updated_count}\n" \
                f"- Services in Excel: {len(excel_services_map)}"
            )
            self.append_log("Completed entering service prices")
        except Exception as e:
            self.logger.error(f"Error entering service prices: {str(e)}")
            raise

    def save_bill_changes(self, page):
        """Save changes to the bill with improved reliability"""
        try:
            self.check_pause()
            self.append_log("Saving bill changes...")

            selectors = [
                'button[id*="save"]',
                'button:has-text("Save")',
                'input[type="button"][value="Save"]',
                'input[type="submit"][value*="Save"]',
                '#toolactions_SAVE-tbb',
                '#toolactions_SAVE-tbb_image'
            ]

            max_attempts = 3
            save_successful = False
            for attempt in range(max_attempts):
                if save_successful:
                    break
                for selector in selectors:
                    try:
                        save_button = page.locator(selector).first
                        if save_button and save_button.is_visible():
                            try:
                                save_button.click(timeout=5000)
                                page.wait_for_load_state("networkidle", timeout=10000)
                                page.wait_for_timeout(500)
                                # Update page reference if context exists
                                if self.context:
                                    page = self.context.pages[-1]
                                save_successful = True
                                self.append_log(f"Clicked save button with selector: {selector}")
                                break
                            except Exception as click_error:
                                self.logger.warning(f"Standard click failed with selector {selector}: {str(click_error)}")
                                try:
                                    page.evaluate(f'document.querySelector("{selector}").click()')
                                    save_successful = True
                                    self.append_log(f"Clicked save button using JavaScript with selector: {selector}")
                                    break
                                except Exception as js_error:
                                    self.logger.warning(f"JavaScript click failed with selector {selector}: {str(js_error)}")
                    except Exception as e:
                        self.logger.warning(f"Error with selector {selector}: {str(e)}")
                        continue
                if not save_successful and attempt < max_attempts - 1:
                    self.logger.warning(f"Save attempt {attempt + 1} failed, retrying...")
                    time.sleep(2)

            if save_successful:
                try:
                    try:
                        save_indicator = page.locator('div[id*="progress"], .loading-indicator').first
                        if save_indicator.is_visible():
                            save_indicator.wait_for(state='hidden', timeout=30000)
                    except Exception:
                        pass
                    page.wait_for_load_state("domcontentloaded", timeout=30000)
                    page.wait_for_load_state("networkidle", timeout=30000)
                    time.sleep(2)
                    success = False
                    for selector in ['text="Save Successful"', 'text="Changes Saved"', '.success-message']:
                        try:
                            indicator = page.locator(selector).first
                            if indicator and indicator.is_visible():
                                success = True
                                break
                        except Exception:
                            continue
                    if success:
                        self.append_log("Save confirmed successful")
                    else:
                        page.wait_for_load_state("networkidle", timeout=15000)
                except Exception as e:
                    self.logger.warning(f"Could not verify save success: {str(e)}")
            else:
                raise Exception("Could not find or click save button after all attempts")
        except Exception as e:
            self.logger.error(f"Error saving bill changes: {str(e)}")
            raise

    def verify_login_success(self, page):
        max_attempts = 2
        attempt_delay = 1
        success_indicators = ['Welcome', 'Logout', 'Sign Out', 'Main Menu']
        for attempt in range(max_attempts):
            try:
                current_url = page.url
                if 'login' not in current_url.lower():
                    self.logger.info("Login successful - Redirected to main page")
                    return True
                success = page.evaluate(f"""
                    (() => {{
                        const allText = document.body.innerText;
                        return {repr(success_indicators)}.some(text => allText.includes(text));
                    }})()
                """)
                if success:
                    self.logger.info("Login successful - Found success indicators")
                    return True
                current_url = page.url
                if 'login' not in current_url.lower():
                    self.logger.info("Login successful - Redirected to main page")
                    return True
                if attempt < max_attempts - 1:
                    self.logger.warning(
                        f"Login verification attempt {attempt + 1} failed, waiting {attempt_delay} seconds..."
                    )
                    time.sleep(attempt_delay)
            except Exception as e:
                self.logger.warning(
                    f"Login verification attempt {attempt + 1} failed: {str(e)}"
                )
                if attempt < max_attempts - 1:
                    time.sleep(attempt_delay)
                continue
        self.logger.error("Login verification failed after maximum attempts")
        return False

    def handle_language_selection(self, page):
        selectors = [
            'a:has-text("English")',
            'a:has-text("EN")',
            'button:has-text("English")',
            'button:has-text("EN")'
        ]
        for sel in selectors:
            try:
                loc = page.locator(sel).first
                if loc and loc.is_visible():
                    loc.click()
                    page.wait_for_load_state("networkidle", timeout=10000)
                    self.append_log("Switched language to English via selector: " + sel)
                    time.sleep(1)
                    break
            except Exception:
                continue

    def perform_login(self, page):
        try:
            self.append_log("Opening login page...", "info")
            page.wait_for_load_state("domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            time.sleep(1)
            page.mouse.move(250, 150)
            try:
                page.wait_for_load_state("networkidle", timeout=15000)
                page.wait_for_load_state("domcontentloaded", timeout=10000)
                success = False
                max_form_attempts = 3
                for attempt in range(max_form_attempts):
                    try:
                        form_present = page.evaluate("""
                            () => {
                                const inputs = document.getElementsByTagName('input');
                                for (let input of inputs) {
                                    if (input.type === 'text' || input.type === 'password' || 
                                        input.name === 'username' || input.name === 'password') {
                                        return true;
                                    }
                                }
                                return false;
                            }
                        """)
                        if form_present:
                            success = True
                            self.append_log("Login form detected successfully", "info")
                            break
                        else:
                            self.append_log(f"Login form not found, attempt {attempt + 1} of {max_form_attempts}", "warning")
                            time.sleep(2)
                    except Exception as e:
                        self.append_log(f"Error detecting form, attempt {attempt + 1}: {str(e)}", "warning")
                        time.sleep(2)
                if not success:
                    raise Exception("Could not detect login form after multiple attempts")
            except Exception as e:
                self.append_log(f"Error while waiting for login form: {str(e)}", "error")
                raise

            self.handle_language_selection(page)
            time.sleep(1)
            max_login_attempts = 3
            for attempt in range(max_login_attempts):
                try:
                    _ = page.evaluate("""
                        () => {
                            const inputs = Array.from(document.getElementsByTagName('input'));
                            return inputs.map(input => ({
                                type: input.type,
                                name: input.name,
                                id: input.id,
                                placeholder: input.placeholder,
                                value: input.value,
                                isVisible: input.offsetParent !== null
                            }));
                        }
                    """)
                    username_field = None
                    visible_text_inputs = [
                        page.locator('input[type="text"]'),
                        page.locator('input:not([type])'),
                        page.locator('input[name="username"]'),
                        page.locator('input[id="username"]'),
                        page.locator('input.loginid'),
                        page.locator('input[name="j_username"]')
                    ]
                    for input_locator in visible_text_inputs:
                        try:
                            if input_locator.is_visible():
                                username_field = input_locator
                                self.logger.info("Found username field")
                                break
                        except Exception:
                            continue
                    if not username_field:
                        try:
                            all_inputs = page.locator('input').all()
                            for input_el in all_inputs:
                                if input_el.is_visible() and not input_el.get_attribute('type') == 'password':
                                    username_field = input_el
                                    self.logger.info("Found username field using general selector")
                                    break
                        except Exception:
                            pass
                    if not username_field:
                        raise Exception("Username field not found after exhaustive search")
                    username_field.hover()
                    time.sleep(random.uniform(0.3, 0.7))
                    username_field.click()
                    username_field.fill("")
                    username = self.username_var.get()
                    username_field.type(username, delay=50)
                    time.sleep(0.2)
                except Exception:
                    self.append_log("Unable to find username field by primary selectors; trying to fill via JS", "warning")
                    page.evaluate(f"document.querySelector('input[name=username]') && (document.querySelector('input[name=username]').value = '{self.username_var.get()}')")

                time.sleep(random.uniform(0.8, 1.5))
                try:
                    password_field = page.locator('input#password, input[name="password"], input[type="password"]').first
                    password_field.hover()
                    time.sleep(random.uniform(0.3, 0.7))
                    password_field.click()
                    password_field.fill("")
                    password = self.password_var.get()
                    for char in password:
                        password_field.type(char, delay=random.uniform(100, 300))
                        time.sleep(random.uniform(0.1, 0.2))
                except Exception:
                    self.append_log("Unable to find password field; attempting JS set", "warning")
                    page.evaluate(f"document.querySelector('input[name=password]') && (document.querySelector('input[name=password]').value = '{self.password_var.get()}')")

                time.sleep(random.uniform(1, 2))
                clicked = False
                selectors = [
                    'button:has-text("Sign In")',
                    'button:has-text("SignIn")',
                    'button:has-text("Login")',
                    'input[type="submit"]',
                    'button[type="submit"]',
                    'input[value="Sign In"]',
                    'input[value="Login"]'
                ]
                for sel in selectors:
                    try:
                        loc = page.locator(sel).first
                        if loc and loc.is_visible():
                            loc.hover()
                            time.sleep(random.uniform(0.3, 0.7))
                            loc.click()
                            clicked = True
                            self.append_log(f"Clicked login with selector: {sel}")
                            break
                    except Exception:
                        continue
                if not clicked:
                    self.append_log("Clicking login using fallback JS", "info")
                    page.evaluate("""
                        (() => {
                            const buttons = Array.from(document.querySelectorAll('input, button'));
                            const loginButton = buttons.find(btn => {
                                const v = (btn.value || btn.textContent || '').toLowerCase();
                                return v.includes('sign in') || v.includes('signin') || v.includes('login');
                            });
                            if (loginButton) { loginButton.click(); return true; }
                            return false;
                        })()
                    """)
                page.wait_for_load_state("networkidle", timeout=30000)
                time.sleep(2)
                if self.verify_login_success(page):
                    self.append_log("✅ Login successful - Welcome to Maximo system!", "info")
                    self.update_status("Login successful", "info")
                    return
                if self.url_var.get().strip().split("?")[0] in page.url:
                    possible_error = ""
                    try:
                        possible_error = page.locator('.error, .validation, .message').first.inner_text()
                    except Exception:
                        pass
                    if possible_error:
                        raise Exception(f"Login failed: {possible_error}")
                if attempt < max_login_attempts - 1:
                    self.append_log("Login attempt unsuccessful, trying again...", "warning")
                    time.sleep(2)
                    continue
        except Exception as e:
            self.logger.exception("perform_login failed")
            raise

    def search_work_order(self, page, wo_number):
        try:
            self.update_status("Opening main menu...", "info")
            try:
                menu_button = page.wait_for_selector('#titlebar-tb_gotoButton', timeout=10000)
                if menu_button:
                    menu_button.click()
                    time.sleep(2)
            except Exception as e:
                self.logger.warning(f"Could not find menu button by ID: {e}")
                try:
                    page.evaluate("""
                        (() => {
                            const buttons = document.querySelectorAll('button');
                            for (const btn of buttons) {
                                const rect = btn.getBoundingClientRect();
                                if (rect.x < 100 && rect.y < 100) {
                                    btn.click();
                                    return true;
                                }
                            }
                            return false;
                        })()
                    """)
                    time.sleep(2)
                except Exception as e:
                    self.logger.error(f"Failed to click menu button: {str(e)}")
                    raise
            self.update_status("Navigating to Work Orders...", "info")
            try:
                work_orders = page.wait_for_selector('#menu0_WO_MODULE_a_tnode', timeout=5000)
                if work_orders:
                    work_orders.click()
                    time.sleep(2)
                else:
                    raise Exception("Work Orders menu item not found")
                tracking_clicked = False
                max_attempts = 3
                for attempt in range(max_attempts):
                    try:
                        selectors = [
                            'text="Work Order Tracking (SP)"',
                            'span:has-text("Work Order Tracking (SP)")',
                            'a:has-text("Work Order Tracking (SP)")',
                            'td:has-text("Work Order Tracking (SP)")',
                            '[title*="Work Order Tracking"]'
                        ]
                        for selector in selectors:
                            try:
                                tracking = page.wait_for_selector(selector, timeout=3000)
                                if tracking and tracking.is_visible():
                                    tracking.click()
                                    self.logger.info(f"Clicked Work Order Tracking using selector: {selector}")
                                    tracking_clicked = True
                                    time.sleep(2)
                                    break
                            except Exception:
                                continue
                        if tracking_clicked:
                            break
                        if not tracking_clicked:
                            found = page.evaluate("""
                                () => {
                                    const elements = Array.from(document.querySelectorAll('*'));
                                    const tracking = elements.find(el => 
                                        el.textContent && 
                                        el.textContent.includes('Work Order Tracking (SP)')
                                    );
                                    if (tracking) {
                                        tracking.click();
                                        return true;
                                    }
                                    return false;
                                }
                            """)
                            if found:
                                self.logger.info("Clicked Work Order Tracking using JavaScript")
                                tracking_clicked = True
                                time.sleep(2)
                                break
                    except Exception as e:
                        self.logger.warning(f"Attempt {attempt + 1} failed: {str(e)}")
                        if attempt < max_attempts - 1:
                            time.sleep(1)
                            continue
                if not tracking_clicked:
                    raise Exception("Could not find or click Work Order Tracking menu item after all attempts")
                page.wait_for_load_state("networkidle", timeout=30000)
                time.sleep(2)
                self.update_status(f"Searching for Work Order {wo_number}...", "info")
                search_box = page.wait_for_selector('input[aria-label="Find Work Order"]', timeout=5000)
                if search_box:
                    search_box.fill('')
                    time.sleep(0.5)
                    search_box.type(str(wo_number), delay=100)
                    time.sleep(1)
                else:
                    raise Exception("Search box not found")
                try:
                    search_btn = page.wait_for_selector('#quicksearchQSImage', timeout=10000)
                    if search_btn:
                        search_btn.click()
                        self.logger.info("Clicked search button")
                    else:
                        raise Exception("Search button not found")
                except Exception as e:
                    self.logger.warning(f"Could not click search button normally: {e}")
                    search_box.press('Enter')
                    self.logger.info("Used Enter key to search")
                page.wait_for_load_state("networkidle", timeout=30000)
                time.sleep(2)
                self.logger.info("Proceeding with next steps...")
                return True
            except Exception as e:
                self.logger.error(f"Navigation error: {e}")
                raise
        except Exception as e:
            self.logger.exception(f"Failed to search for Work Order {wo_number}")
            raise Exception(f"Search failed: {str(e)}")

    def process_customer_bill(self, page):
        """Edit existing bill prices across Services pages and save changes (parity path)."""
        try:
            self.check_pause()
            self.append_log("Starting to update service prices from Excel (Services tab)...")
            # enter_services_prices already clicks Services tab and iterates pages
            self.enter_services_prices(page)
            # Save changes
            self.save_bill_changes(page)
        except Exception as e:
            self.logger.error(f"Error processing customer bill: {str(e)}")
            raise

    def update_excel_description(self, description):
        """Update description in memory and persist to Excel (B3)."""
        try:
            self.wo_description = description
            self.save_description_to_excel_file(description)
            self.append_log(f"Updated Excel description: {description}")
        except Exception:
            self.logger.exception("update_excel_description failed")

    def save_description_to_excel_file(self, description):
        """Save description directly to Excel file (B3 cell in current sheet)."""
        try:
            if not hasattr(self, 'working_excel_path') or not self.working_excel_path:
                self.logger.warning("No Excel working copy path available")
                return False

            from openpyxl import load_workbook
            is_xlsm = str(Path(self.working_excel_path).suffix).lower() == ".xlsm"
            wb = load_workbook(self.working_excel_path, keep_vba=is_xlsm)

            if hasattr(self, 'current_sheet') and self.current_sheet in wb.sheetnames:
                ws = wb[self.current_sheet]
            else:
                ws = wb.active

            ws['B3'] = description
            wb.save(self.working_excel_path)
            wb.close()
            self.append_log(f"✅ Description saved to Excel (Sheet: {ws.title}, B3): {description}")
            return True
        except Exception as e:
            self.logger.exception("Failed to save description to Excel file")
            self.append_log(f"❌ Error saving description to Excel: {str(e)}", "error")
            return False

    def navigate_to_tab(self, page, tab_name):
        """Click tab with given text and wait."""
        try:
            self.append_log(f"Switching to tab: {tab_name}")
            tab = page.locator(f'a:has-text("{tab_name}"), button:has-text("{tab_name}")').first
            tab.click()
            page.wait_for_load_state("networkidle", timeout=15000)
            time.sleep(0.5)
        except Exception:
            self.append_log(f"Could not click tab {tab_name} via primary selectors; trying fallback", "warning")
            try:
                page.evaluate(
                    """
                    (name) => {
                        const el = Array.from(document.querySelectorAll('*')).find(n => n.textContent && n.textContent.trim() === name);
                        if (el) el.click();
                    }
                    """,
                    tab_name,
                )
                page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                self.append_log(f"Fallback click for tab {tab_name} failed", "error")

    def add_services_to_wo(self, page):
        """Add services listed in self.services_data to the current Work Order."""
        try:
            self.append_log("Navigating to Actuals tab...")
            actuals_tab = page.locator('#m272f5640-tab_anchor').first
            actuals_tab.click()
            page.wait_for_load_state("networkidle", timeout=30000)
            time.sleep(2)

            self.append_log("Navigating to Services tab...")
            services_tab = page.locator('#m74d804c0-tab_anchor').first
            services_tab.click()
            page.wait_for_load_state("networkidle", timeout=30000)

            if hasattr(self, 'current_sheet') and self.current_sheet != 'Sheet1':
                time.sleep(3)
            else:
                time.sleep(2)

            if not hasattr(self, 'services_data') or self.services_data is None or self.services_data.empty:
                self.append_log("No services data found to process")
                return

            total_services = len(self.services_data)
            processed_services = 0
            for idx, (_, service) in enumerate(self.services_data.iterrows()):
                if self.stop_event.is_set():
                    self.append_log("Stop requested while adding services.")
                    break

                row_num = idx + 8
                quantity = service.get('Quantity', 0)
                service_item = service.get('Service Item', '')

                if quantity == 0 or pd.isna(quantity):
                    self.append_log(f"⏭️ SKIPPING service '{service_item}' from Excel row {row_num} - quantity is ZERO: {quantity}")
                    continue

                if not service_item or str(service_item).strip() == '':
                    self.append_log(f"⚠️ SKIPPING service from Excel row {row_num} - service item is EMPTY")
                    continue

                if quantity == 1:
                    self.append_log(f"📦 Processing service: '{service_item}' with DEFAULT quantity: {quantity} (won't change)")
                else:
                    self.append_log(f"📦 Processing service: '{service_item}' with CUSTOM quantity: {quantity} (will set)")

                service_dict = {
                    'Service Item': service.get('Service Item'),
                    'Description': service.get('Description'),
                    'Quantity': service.get('Quantity'),
                    'Unit Price': service.get('Unit Price'),
                    'Total Price': service.get('Total Price')
                }

                self.append_log(f"Processing service from Excel row {row_num}...")
                self.add_single_service(page, service_dict, row_num)
                processed_services += 1

                if total_services:
                    # Update inner-sheet progress as a hint (optional)
                    set_progress(min(0.99, processed_services / max(1, total_services)))

            self.route_wo_to_comp(page)
        except Exception:
            self.logger.exception("add_services_to_wo failed")
            raise

    def add_single_service(self, page, service, excel_row):
        """Add a single service row with enhanced error handling and retry logic."""
        start_time = time.time()
        max_total_time = 120
        try:
            service_item = service.get('Service Item', '')
            quantity = service.get('Quantity', 1)
            if not service_item or str(service_item).strip() == '':
                raise Exception("Service Item is empty or invalid")
            if service_item == "IS00000082":
                self.append_log(f"🔧 Detected problematic service {service_item} - using enhanced handling")
                time.sleep(1)

            if quantity == 1:
                self.append_log(f"⚙️ Adding service: {service_item} (Quantity: {quantity} - DEFAULT, no input needed)")
            else:
                self.append_log(f"⚙️ Adding service: {service_item} (Quantity: {quantity} - CUSTOM, will input)")
            self.logger.info(f"Adding service {service_item} from Excel row {excel_row} with quantity {quantity}")

            if hasattr(self, 'stop_event') and self.stop_event.is_set():
                raise Exception("Operation was stopped by user")

            def check_timeout():
                self.check_pause()
                if time.time() - start_time > max_total_time:
                    raise Exception(f"Service addition timeout after {max_total_time} seconds for {service_item}")
                if hasattr(self, 'stop_event') and self.stop_event.is_set():
                    raise Exception("Operation was stopped by user")

            try:
                check_timeout()
                self.append_log(f"🔘 Clicking 'Select Internal Service' button...")
                select_service_btn = page.locator('#md5cf4765_bg_button_pluspselectinternalservice-pb').first
                select_service_btn.wait_for(state='visible', timeout=8000)
                try:
                    page.wait_for_selector('.wait_modal, .loading-overlay, .loading-indicator', state='hidden', timeout=5000)
                except:
                    pass
                time.sleep(0.5)
                select_service_btn.click()
                self.append_log(f"✅ Successfully clicked 'Select Internal Service'")
                time.sleep(0.5)
                check_timeout()
            except Exception as e:
                self.logger.error(f"Failed to click Select Internal Service button: {e}")
                raise

            try:
                check_timeout()
                self.append_log(f"🔍 Opening filter dialog...")
                filter_btn = page.locator('#meb6466f-ti_img').first
                filter_btn.wait_for(state='visible', timeout=8000)
                try:
                    page.wait_for_selector('.wait_modal, .loading-overlay, .loading-indicator', state='hidden', timeout=5000)
                except:
                    pass
                time.sleep(0.5)
                filter_btn.click()
                self.append_log(f"✅ Filter dialog opened")
                time.sleep(1.0)
                check_timeout()
            except Exception as e:
                self.logger.error(f"Failed to click Filter button: {e}")
                raise

            max_search_attempts = 3
            search_successful = False
            for attempt in range(max_search_attempts):
                self.check_pause()
                try:
                    self.append_log(f"🔍 Search attempt {attempt + 1}/{max_search_attempts} for service: {service_item}")
                    search_field = None
                    selectors = [
                        'input[aria-labelledby="meb6466f_ttrow_[C:2]-c"]',
                        '[role="textbox"][aria-labelledby="meb6466f_ttrow_[C:2]-c"]',
                        'input.queryField[type="text"]',
                        'input[id*="meb6466f"][type="text"]',
                        '.queryField',
                        'input[placeholder*="service"]',
                        'input[placeholder*="Service"]',
                        'input[placeholder*="type"]',
                        '.tt_content input[type="text"]',
                        'div[id*="filter"] input[type="text"]',
                        'table[role="presentation"] input[type="text"]',
                        'input[type="text"]:visible'
                    ]
                    for selector in selectors:
                        self.check_pause()
                        try:
                            element = page.locator(selector).first
                            element.wait_for(state='visible', timeout=3000)
                            if element.is_visible():
                                search_field = element
                                self.logger.info(f"Found search field with selector: {selector}")
                                break
                        except Exception as selector_error:
                            self.logger.debug(f"Selector {selector} failed: {selector_error}")
                            continue
                    if not search_field:
                        if attempt < max_search_attempts - 1:
                            self.append_log(f"⚠️ Search field not found, retrying in 2 seconds...")
                            time.sleep(2)
                            continue
                        else:
                            raise Exception("Could not find Service Item search field after all attempts")

                    self.append_log(f"📝 Entering service item: {service_item}")
                    search_field.click(); time.sleep(0.3)
                    try:
                        search_field.press("Control+a"); time.sleep(0.1)
                        search_field.press("Delete"); time.sleep(0.1)
                        search_field.clear()
                    except:
                        pass
                    clean_service_item = str(service_item).strip().upper()
                    input_successful = False
                    input_methods = [
                        lambda: search_field.fill(clean_service_item),
                        lambda: search_field.type(clean_service_item, delay=50),
                        lambda: page.evaluate(f"document.querySelector('{selectors[0]}').value = '{clean_service_item}'")
                    ]
                    for method in input_methods:
                        self.check_pause()
                        try:
                            method(); time.sleep(0.2)
                            current_value = search_field.input_value()
                            if current_value.strip().upper() == clean_service_item:
                                input_successful = True
                                break
                        except Exception as input_error:
                            self.logger.debug(f"Input method failed: {input_error}")
                            continue
                    if not input_successful:
                        if attempt < max_search_attempts - 1:
                            self.append_log(f"⚠️ Failed to enter service item, retrying...")
                            continue
                        else:
                            raise Exception(f"Failed to enter service item: {clean_service_item}")
                    search_field.press('Enter')
                    self.append_log(f"⏳ Searching for service {clean_service_item}...")
                    time.sleep(1.5)
                    results_indicators = [
                        '#meb6466f_tdrow_\\[C\\:0\\]_tbselrow-ti\\[R\\:0\\]_img',
                        'img[id*="meb6466f_tdrow"][id*="tbselrow"]',
                        '[id*="meb6466f_tdrow"]',
                        'tr[id*="meb6466f_tdrow"]',
                        '.tablecontent tr[id*="meb6466f_tdrow"]',
                        '.tablecontent tbody tr',
                        'td:has-text("' + clean_service_item + '")',
                        'span:has-text("' + clean_service_item + '")',
                        'table[role="presentation"] tbody tr',
                        '.tt_content tbody tr'
                    ]
                    search_result_found = False
                    for i, indicator in enumerate(results_indicators):
                        self.check_pause()
                        try:
                            element = page.locator(indicator).first
                            element.wait_for(state='visible', timeout=3000)
                            if element.is_visible():
                                search_result_found = True
                                self.append_log(f"✅ Found search results using strategy {i+1}: {indicator}")
                                break
                        except Exception as indicator_error:
                            self.logger.debug(f"Indicator {indicator} failed: {indicator_error}")
                            continue
                    if not search_result_found:
                        try:
                            row_count = page.locator('tbody tr').count()
                            if row_count > 0:
                                search_result_found = True
                                self.append_log(f"✅ Found {row_count} result rows")
                        except:
                            pass
                    if search_result_found:
                        search_successful = True
                        break
                    else:
                        if attempt < max_search_attempts - 1:
                            self.append_log(f"⚠️ No results found for {clean_service_item}, retrying search...")
                            continue
                        else:
                            self.append_log(f"❌ Service {clean_service_item} not found after all attempts")
                            raise Exception(f"Service {clean_service_item} not found in search results")
                except Exception as search_error:
                    if attempt < max_search_attempts - 1:
                        self.append_log(f"⚠️ Search attempt {attempt + 1} failed: {search_error}")
                        time.sleep(2)
                        continue
                    else:
                        self.logger.error(f"All search attempts failed for {service_item}: {search_error}")
                        raise Exception(f"Failed to search for service {service_item} after {max_search_attempts} attempts: {search_error}")
            if not search_successful:
                raise Exception(f"Failed to complete search for service {service_item}")
            self.append_log(f"✅ Successfully found service {service_item}")

            checkbox_clicked = False
            max_checkbox_attempts = 5
            for checkbox_attempt in range(max_checkbox_attempts):
                self.check_pause()
                try:
                    self.append_log(f"☑️ Selecting service (attempt {checkbox_attempt + 1}/{max_checkbox_attempts})...")
                    time.sleep(1)
                    checkbox_selectors = [
                        '#meb6466f_tdrow_\\[C\\:0\\]_tbselrow-ti\\[R\\:0\\]_img',
                        'img[id*="meb6466f_tdrow_"][id*="_tbselrow-ti"][id*="R:0"]',
                        '[id*="meb6466f_tdrow"][id*="tbselrow-ti"][id*="_img"]',
                        'table[role="presentation"] tbody tr:first-child td:first-child img',
                        'table[role="presentation"] tr[id*="meb6466f_tdrow"] td:first-child img',
                        '[id*="meb6466f_tdrow"][id*="[C:0]"] img',
                        'td[id*="meb6466f_tdrow_[C:0]"] img',
                        '.tablecontent tr:first-child td:first-child img',
                        '.tablecontent img[title*="Select"]',
                        '.tablecontent input[type="checkbox"]'
                    ]
                    checkbox = None
                    working_selector = None
                    for selector in checkbox_selectors:
                        self.check_pause()
                        try:
                            element = page.locator(selector).first
                            element.wait_for(state='visible', timeout=4000)
                            if element.is_visible():
                                checkbox = element; working_selector = selector
                                self.logger.info(f"Found checkbox with selector: {selector}")
                                break
                        except Exception as selector_error:
                            self.logger.debug(f"Selector {selector} failed: {selector_error}")
                            continue
                    if not checkbox:
                        self.append_log(f"⚠️ Direct checkbox not found, trying to click first table row...")
                        try:
                            first_row = page.locator('tr[id*="meb6466f_tdrow"]').first
                            first_row.wait_for(state='visible', timeout=3000)
                            if first_row.is_visible():
                                first_row.click(); checkbox_clicked = True
                                self.append_log(f"✅ Selected service by clicking table row")
                                break
                        except Exception as row_error:
                            self.logger.debug(f"Table row click failed: {row_error}")
                    if not checkbox and checkbox_attempt < max_checkbox_attempts - 1:
                        self.append_log(f"⚠️ Checkbox not found, retrying in 2 seconds...")
                        time.sleep(2)
                        continue
                    elif not checkbox:
                        try:
                            self.append_log(f"🔧 Final attempt: Using JavaScript to find and click checkbox...")
                            js_result = page.evaluate("""
                                () => {
                                    const images = document.querySelectorAll('img[id*="tbselrow"]');
                                    if (images.length > 0) { images[0].click(); return 'clicked_image'; }
                                    const rows = document.querySelectorAll('tr[id*="meb6466f_tdrow"]');
                                    if (rows.length > 0) { rows[0].click(); return 'clicked_row'; }
                                    return 'no_element_found';
                                }
                            """)
                            if (js_result != 'no_element_found'):
                                checkbox_clicked = True
                                self.append_log(f"✅ JavaScript click successful: {js_result}")
                                break
                            else:
                                raise Exception("No checkbox or row found via JavaScript")
                        except Exception as js_error:
                            raise Exception(f"All checkbox selection methods failed: {js_error}")
                    try:
                        self.append_log(f"🖱️ Clicking checkbox with selector: {working_selector}")
                        checkbox.click(); time.sleep(0.8)
                        try:
                            verification_passed = False
                            if checkbox.is_visible():
                                verification_passed = True
                            try:
                                selected_row = page.locator('tr[id*="meb6466f_tdrow"].selected').first
                                if selected_row.is_visible():
                                    verification_passed = True
                            except:
                                pass
                            if verification_passed:
                                checkbox_clicked = True
                                self.append_log(f"✅ Service selected successfully")
                                break
                            else:
                                self.append_log(f"⚠️ Selection verification unclear, assuming success")
                                checkbox_clicked = True
                                break
                        except Exception as verify_error:
                            self.logger.debug(f"Verification failed but continuing: {verify_error}")
                            checkbox_clicked = True
                            break
                    except Exception as click_error:
                        if checkbox_attempt < max_checkbox_attempts - 1:
                            self.append_log(f"⚠️ Click failed: {click_error}, retrying...")
                            time.sleep(1.5)
                            continue
                        else:
                            raise Exception(f"Failed to click checkbox: {click_error}")
                except Exception as checkbox_error:
                    if checkbox_attempt < max_checkbox_attempts - 1:
                        self.append_log(f"⚠️ Checkbox attempt {checkbox_attempt + 1} failed: {checkbox_error}")
                        time.sleep(2)
                        continue
                    else:
                        raise Exception(f"Failed to select service checkbox after {max_checkbox_attempts} attempts: {checkbox_error}")
            if not checkbox_clicked:
                raise Exception("Failed to select service checkbox after all attempts")

            self.append_log(f"✅ Service selected, now clicking OK button...")
            ok_clicked = False
            max_ok_attempts = 5
            for ok_attempt in range(max_ok_attempts):
                self.check_pause()
                try:
                    self.append_log(f"🔘 Clicking OK button (attempt {ok_attempt + 1}/{max_ok_attempts})...")
                    ok_selectors = [
                        '#mb1ea0e57-pb',
                        'input[value="OK"][type="button"]',
                        'button:has-text("OK")',
                        'input[title="OK"]',
                        '[role="button"]:has-text("OK")',
                        'td:has-text("OK") input',
                        '.dijitDialog input[type="button"]',
                        '.dialog-buttons input[type="button"]',
                        'input[type="button"][value*="OK"]',
                        'button[title*="OK"]'
                    ]
                    ok_button = None; working_ok_selector = None
                    for selector in ok_selectors:
                        self.check_pause()
                        try:
                            element = page.locator(selector).first
                            element.wait_for(state='visible', timeout=3000)
                            if element.is_visible():
                                ok_button = element; working_ok_selector = selector
                                self.logger.info(f"Found OK button with selector: {selector}")
                                break
                        except Exception as selector_error:
                            self.logger.debug(f"OK selector {selector} failed: {selector_error}")
                            continue
                    if not ok_button:
                        if ok_attempt < max_ok_attempts - 1:
                            self.append_log(f"⚠️ OK button not found, retrying in 1 second...")
                            time.sleep(1)
                            continue
                        else:
                            try:
                                self.append_log(f"🔧 Final attempt: Using JavaScript to find and click OK button...")
                                js_result = page.evaluate("""
                                    () => {
                                        const okInputs = document.querySelectorAll('input[type="button"][value*="OK"]');
                                        if (okInputs.length > 0) { okInputs[0].click(); return 'clicked_input_ok'; }
                                        const okButtons = document.querySelectorAll('button');
                                        for (let btn of okButtons) { if (btn.textContent.includes('OK')) { btn.click(); return 'clicked_button_ok'; } }
                                        const okElement = document.querySelector('[id*="mb1ea0e57"]');
                                        if (okElement) { okElement.click(); return 'clicked_id_element'; }
                                        return 'no_ok_button_found';
                                    }
                                """)
                                if js_result != 'no_ok_button_found':
                                    ok_clicked = True
                                    self.append_log(f"✅ JavaScript OK click successful: {js_result}")
                                    break
                                else:
                                    raise Exception("No OK button found via JavaScript")
                            except Exception as js_error:
                                raise Exception(f"All OK button methods failed: {js_error}")
                    try:
                        self.append_log(f"🖱️ Clicking OK button with selector: {working_ok_selector}")
                        time.sleep(0.5)
                        ok_button.click()
                        time.sleep(2)
                        try:
                            dialog_visible = page.locator('.dijitDialog, [role="dialog"]').first.is_visible()
                            if not dialog_visible:
                                ok_clicked = True
                                self.append_log(f"✅ OK button clicked successfully - dialog closed")
                                break
                            else:
                                self.append_log(f"⚠️ Dialog still visible after OK click, assuming success")
                                ok_clicked = True
                                break
                        except:
                            ok_clicked = True
                            self.append_log(f"✅ OK button clicked - assuming success")
                            break
                    except Exception as click_error:
                        if ok_attempt < max_ok_attempts - 1:
                            self.append_log(f"⚠️ OK click failed: {click_error}, retrying...")
                            time.sleep(1.5)
                            continue
                        else:
                            raise Exception(f"Failed to click OK button: {click_error}")
                except Exception as ok_error:
                    if ok_attempt < max_ok_attempts - 1:
                        self.append_log(f"⚠️ OK attempt {ok_attempt + 1} failed: {ok_error}")
                        time.sleep(2)
                        continue
                    else:
                        raise Exception(f"Failed to click OK button after {max_ok_attempts} attempts: {ok_error}")
            if not ok_clicked:
                raise Exception("Failed to click OK button after all attempts")

            try:
                quantity = float(service.get('Quantity', 1.00))
                if quantity == 1.0:
                    self.logger.info(f"📊 Quantity is 1.0 (default) - NO QUANTITY UPDATE NEEDED")
                    self.append_log(f"📊 Quantity is 1 (default) - skipping quantity input completely")
                else:
                    self.logger.info(f"📊 Quantity is {quantity} (custom) - UPDATING quantity field")
                    self.append_log(f"📊 Setting custom quantity: {quantity}")
                    time.sleep(2)
                    quantity_selectors = [
                        '#md5cf4765_tdrow_\\[C\\:4\\]_txt-tb\\[R\\:0\\]',
                        'input[id*="_tdrow_"][id*="[C:4]"][role="textbox"]',
                        'input[id*="quantity"][role="textbox"]'
                    ]
                    quantity_field = None
                    for selector in quantity_selectors:
                        try:
                            field = page.wait_for_selector(selector, timeout=10000)
                            if field and field.is_visible():
                                quantity_field = field
                                self.logger.info(f"Found quantity field using selector: {selector}")
                                break
                        except Exception as e:
                            self.logger.warning(f"Failed to find field with selector {selector}: {str(e)}")
                            continue
                    if not quantity_field:
                        raise Exception("Could not find quantity field with any known selector")
                    self.logger.info(f"Found quantity field, updating to {quantity}")
                    max_attempts = 3
                    for attempt in range(max_attempts):
                        self.check_pause()
                        try:
                            time.sleep(0.5)
                            quantity_field.click()
                            quantity_field.press("Control+a")
                            quantity_field.press("Delete")
                            page.wait_for_timeout(500)
                            formatted_quantity = "{:.2f}".format(float(quantity))
                            quantity_field.fill(formatted_quantity)
                            page.wait_for_timeout(500)
                            quantity_field.press('Tab')
                            page.wait_for_timeout(1000)
                            actual_value = quantity_field.input_value().strip()
                            expected_value = formatted_quantity
                            if actual_value == expected_value:
                                self.logger.info(f"Successfully set quantity to {formatted_quantity}")
                                break
                            else:
                                self.logger.warning(f"Attempt {attempt + 1}: Value mismatch - Got '{actual_value}', Expected '{expected_value}'")
                                if attempt < max_attempts - 1:
                                    page.wait_for_timeout(1000)
                                else:
                                    raise Exception(f"Failed to set quantity after {max_attempts} attempts")
                        except Exception as e:
                            self.logger.error(f"Error updating quantity on attempt {attempt + 1}: {str(e)}")
                            if attempt == max_attempts - 1:
                                raise Exception(f"Failed to update quantity after {max_attempts} attempts: {str(e)}")
                            page.wait_for_timeout(1000)
                try:
                    self.logger.info("Attempting to click save button...")
                    save_button = page.locator('#toolactions_SAVE-tbb_image').first
                    save_button.wait_for(state='visible', timeout=5000)
                    save_button.click(timeout=5000)
                    self.logger.info("Successfully clicked save button")
                    page.wait_for_load_state("networkidle", timeout=10000)
                    time.sleep(2)
                    self.logger.info("Save operation completed successfully")
                except Exception as save_error:
                    self.logger.error(f"Failed to save: {save_error}")
                    raise
            except Exception as e:
                self.logger.error(f"Failed to update quantity: {e}")
                raise
            self.logger.info(f"Successfully added service item: {service_item} with quantity: {quantity}")
            return True
        except Exception as e:
            self.logger.error(f"Error adding service item: {e}")
            return False

    def click_route_button(self, page):
        """Helper function to click the route button with multiple fallback methods"""
        try:
            try:
                route_btn = page.locator('#ROUTEWF_IBMMAINWF_-tbb_image').first
                if route_btn and route_btn.is_visible():
                    route_btn.click()
                    page.wait_for_load_state("networkidle", timeout=10000)
                    self.logger.info("Clicked Route button by specific ID")
                    return True
            except Exception as e:
                self.logger.warning(f"Could not find Route button by ID: {e}")
            route_selectors = [
                'img[alt="Route Workflow"]',
                'button:has-text("Route")',
                'a:has-text("Route")',
                'img[src*="route.gif"]',
                '#ROUTEWF_-tbb_image',
                '[aria-label*="Route"]'
            ]
            for selector in route_selectors:
                try:
                    btn = page.locator(selector).first
                    if btn and btn.is_visible():
                        btn.hover(); time.sleep(0.2)
                        btn.click()
                        page.wait_for_load_state("networkidle", timeout=10000)
                        self.logger.info(f"Clicked Route button using selector: {selector}")
                        return True
                except Exception:
                    continue
            self.logger.warning("Trying JavaScript to find and click Route button")
            found = page.evaluate("""
                (() => {
                    let btn = document.querySelector('#ROUTEWF_IBMMAINWF_-tbb_image, #ROUTEWF_-tbb_image');
                    if (!btn) {
                        btn = Array.from(document.querySelectorAll('img,button,a,input')).find(el => {
                            const text = (el.textContent || '').toLowerCase();
                            const alt = (el.alt || '').toLowerCase();
                            const src = (el.src || '').toLowerCase();
                            const label = (el.getAttribute('aria-label') || '').toLowerCase();
                            return alt.includes('route workflow') || text.includes('route') || src.includes('route.gif') || label.includes('route');
                        });
                    }
                    if (btn) { btn.click(); return true; }
                    return false;
                })()
            """)
            if found:
                page.wait_for_load_state("networkidle", timeout=10000)
                self.logger.info("Clicked Route button using JavaScript")
                return True
            self.logger.warning("Could not find Route button with any method")
            return False
        except Exception as e:
            self.logger.warning(f"Failed to click route button: {e}")
            return False

    def route_wo_to_comp(self, page):
        """Route Work Order to COMP status following specific sequence of steps"""
        try:
            self.append_log("🔄 Routing Work Order to COMP (Complete) status...")
            self.append_log("📋 Step 1: Clicking first Route button")
            self.logger.info("Step 1: First Route click")
            if not self.click_route_button(page):
                raise Exception("Failed to click first Route button")
            time.sleep(2)
            self.logger.info("Step 2: Clicking first OK")
            try:
                time.sleep(2)
                ok_clicked = False
                try:
                    ok_btn = page.locator('#m37917b04-pb').first
                    if ok_btn.is_visible(timeout=3000):
                        page.wait_for_selector('#m37917b04-pb', state='visible', timeout=3000)
                        ok_btn.click()
                        ok_clicked = True
                        self.logger.info("Clicked OK button by specific ID")
                except Exception:
                    self.logger.info("Could not find OK button by specific ID")
                if not ok_clicked:
                    try:
                        ok_btn = page.locator('button.text.pb.default:has-text("OK")').first
                        if ok_btn.is_visible(timeout=5000):
                            ok_btn.click(); ok_clicked = True
                            self.logger.info("Clicked OK button by class and text")
                    except Exception:
                        self.logger.info("Could not find OK button by class and text")
                if not ok_clicked:
                    try:
                        ok_btn = page.locator('button:has-text("OK")').first
                        if ok_btn.is_visible(timeout=5000):
                            ok_btn.click(); ok_clicked = True
                            self.logger.info("Clicked OK button by text")
                    except Exception:
                        self.logger.info("Could not find OK button by text")
                if not ok_clicked:
                    try:
                        ok_btn = page.locator('//button[contains(@class, "pb") and contains(text(), "OK")]').first
                        if ok_btn.is_visible(timeout=5000):
                            ok_btn.click(); ok_clicked = True
                            self.logger.info("Clicked OK button by xpath")
                    except Exception:
                        self.logger.info("Could not find OK button by xpath")
                if not ok_clicked:
                    try:
                        ok_clicked = page.evaluate("""
                            () => {
                                const okButton = Array.from(document.querySelectorAll('button')).find(b => b.textContent.includes('OK'));
                                if (okButton) { okButton.click(); return true; }
                                return false;
                            }
                        """)
                    except Exception:
                        self.logger.warning("JavaScript OK click failed")
                if not ok_clicked:
                    raise Exception("Could not click first OK button")
                time.sleep(2)
            except Exception as e:
                self.logger.error(f"Error on first OK: {str(e)}")
                raise
            self.logger.info("Step 3: Second Route click")
            if not self.click_route_button(page):
                raise Exception("Failed to click second Route button")
            time.sleep(2)
            self.logger.info("Step 4: Clicking second OK")
            try:
                time.sleep(2)
                ok_clicked = False
                try:
                    ok_btn = page.locator('#md875c1f9-pb').first
                    if ok_btn.is_visible(timeout=3000):
                        time.sleep(1); ok_btn.click(force=True); ok_clicked = True
                        self.logger.info("Clicked second OK button successfully")
                except Exception as e:
                    self.logger.warning(f"Could not click second OK by ID: {str(e)}")
                if not ok_clicked:
                    try:
                        ok_btn = page.locator('button.text.pb.default:has-text("OK")').first
                        if ok_btn.is_visible(timeout=3000):
                            time.sleep(1); ok_btn.click(force=True); ok_clicked = True
                            self.logger.info("Clicked second OK using class selector")
                    except Exception as e:
                        self.logger.warning(f"Could not click second OK using class selector: {str(e)}")
                if not ok_clicked:
                    try:
                        ok_clicked = page.evaluate("""
                            () => {
                                const okBtn = document.querySelector('#md875c1f9-pb');
                                if (okBtn) { okBtn.click(); return true; }
                                return false;
                            }
                        """)
                        if ok_clicked:
                            self.logger.info("Clicked second OK using JavaScript")
                    except Exception as e:
                        self.logger.error(f"JavaScript click failed: {str(e)}")
                if not ok_clicked:
                    raise Exception("Could not click second OK button with any method")
                time.sleep(2)
            except Exception as e:
                self.logger.error(f"Error on second OK: {str(e)}")
                raise
            self.logger.info("Step 5: Third Route click")
            if not self.click_route_button(page):
                raise Exception("Failed to click third Route button")
            time.sleep(2)
            self.logger.info("Step 6: Selecting Complete Work Order")
            try:
                page.wait_for_selector("input[type='radio']"); time.sleep(1)
                complete_clicked = False
                try:
                    radio = page.locator("label:has-text('Complete Work Order')")
                    if radio.is_visible(timeout=5000):
                        radio.click(); complete_clicked = True
                        self.logger.info("Selected Complete Work Order by label")
                except Exception:
                    self.logger.info("Could not click Complete Work Order by label")
                if not complete_clicked:
                    try:
                        radio = page.locator("input[type='radio']:near(:text('Complete Work Order'))").first
                        if radio.is_visible(timeout=5000):
                            radio.click(); complete_clicked = True
                            self.logger.info("Selected Complete Work Order by radio near text")
                    except Exception:
                        self.logger.info("Could not click radio near Complete Work Order text")
                if not complete_clicked:
                    try:
                        complete_clicked = page.evaluate("""
                            () => {
                                const labels = Array.from(document.querySelectorAll('label'));
                                const targetLabel = labels.find(l => l.textContent.includes('Complete Work Order'));
                                if (targetLabel) {
                                    const radio = document.querySelector(`#${targetLabel.getAttribute('for')}`);
                                    if (radio) { radio.click(); return true; }
                                }
                                return false;
                            }
                        """)
                        if complete_clicked:
                            self.logger.info("Selected Complete Work Order using JavaScript")
                    except Exception:
                        self.logger.warning("JavaScript Complete Work Order selection failed")
                if not complete_clicked:
                    raise Exception("Could not select Complete Work Order option")
                time.sleep(1)
                try:
                    ok_btn = page.locator("button:has-text('OK')").first
                    if ok_btn.is_visible(timeout=5000):
                        ok_btn.click(); self.logger.info("Clicked final OK button")
                except Exception as e:
                    self.logger.warning(f"Could not click final OK button normally: {e}")
                    page.evaluate("""
                        () => {
                            const okButton = Array.from(document.querySelectorAll('button')).find(b => b.textContent.includes('OK'));
                            if (okButton) okButton.click();
                        }
                    """)
                self.logger.info("Proceeding to bill creation without waiting for COMP status")
                return True
            except Exception as e:
                self.logger.error(f"Error in final route step: {str(e)}")
                return True
        except Exception as e:
            self.logger.exception("route_wo_to_comp failed")
            raise

    def save_bill_info(self, page):
        """Save bill information back to Excel"""
        try:
            self.append_log("💾 Saving bill information to Excel file...")
            if not hasattr(self, 'bill_batch') or not self.bill_batch:
                self.logger.warning("No bill_batch found to save")
                return
            if not hasattr(self, 'wo_description') or not self.wo_description or "To be stored" in self.wo_description or "سيتم التخزين" in self.wo_description:
                if hasattr(self, 'excel_original_description') and self.excel_original_description:
                    self.wo_description = self.excel_original_description
                    self.append_log(f"Using Excel original description: {self.wo_description}")
            self.update_excel_file()
            self.append_log(f"Bill information saved - Batch: {self.bill_batch}")
            self.append_log(f"Description saved: {self.wo_description}")
        except Exception as e:
            self.logger.exception("save_bill_info failed")
            raise

    def create_customer_bill(self, page):
        try:
            self.append_log("Creating Customer Bill...")
            page.wait_for_load_state("networkidle", timeout=5000); time.sleep(1)
            try:
                action_input = page.locator('#toolbar2_tbs_1_tbcb_0_action-tb').first
                if action_input.is_visible():
                    action_input.click(); self.append_log("Clicked Select Action input")
                else:
                    dropdown = page.locator('#toolbar2_tbs_1_tbcb_0_action-img').first
                    if dropdown.is_visible():
                        dropdown.click(); self.append_log("Clicked Select Action dropdown arrow")
                    else:
                        raise Exception("Could not find Select Action dropdown")
                time.sleep(1)
                billing_option = page.locator('span[id="menu0_20215C39_HEADER_a_tnode"]').first
                if billing_option.is_visible():
                    billing_option.click(); self.append_log("Selected Customer Billing option"); time.sleep(1)
                else:
                    self.logger.warning("Customer Billing span not visible, trying alternative methods")
                    for selector in ['span:has-text("Customer Billing")','td:has-text("Customer Billing")','div:has-text("Customer Billing")']:
                        try:
                            element = page.locator(selector).first
                            if element.is_visible():
                                element.click(); self.append_log(f"Selected Customer Billing using alternate selector: {selector}"); time.sleep(1)
                                break
                        except Exception as e:
                            self.logger.warning(f"Failed with selector {selector}: {str(e)}")
                            continue
                    else:
                        self.logger.error("Failed to find Customer Billing option with any selector")
                self.append_log("Looking for Create Customer Bill option...")
                create_bill = page.locator('span:has-text("Create Customer Bill")').first
                max_attempts = 5
                for attempt in range(max_attempts):
                    try:
                        if create_bill.is_visible():
                            create_bill.click(); self.append_log("Successfully clicked Create Customer Bill")
                            break
                        else:
                            self.logger.warning(f"Create Customer Bill not visible yet, attempt {attempt + 1}/{max_attempts}")
                            time.sleep(1)
                    except Exception as e:
                        self.logger.warning(f"Attempt {attempt + 1} failed: {str(e)}"); time.sleep(1)
                else:
                    self.logger.error("Failed to click Create Customer Bill after multiple attempts")
                    raise Exception("Could not find Create Customer Bill option")
                page.wait_for_load_state("networkidle", timeout=30000); time.sleep(2)
                try:
                    time.sleep(2)
                    bill_batch_el = page.locator('input[id="m9a223887-tb"]').first
                    if bill_batch_el and bill_batch_el.is_visible():
                        bill_batch_value = bill_batch_el.input_value()
                        if bill_batch_value and bill_batch_value.strip():
                            self.bill_batch = bill_batch_value.strip(); self.append_log(f"Retrieved bill batch number: {self.bill_batch}")
                        else:
                            bill_batch_value = page.evaluate('document.getElementById("m9a223887-tb").value')
                            if bill_batch_value:
                                self.bill_batch = bill_batch_value.strip(); self.append_log(f"Retrieved bill batch via JavaScript: {self.bill_batch}")
                            else:
                                raise Exception("Bill batch field is empty")
                    else:
                        raise Exception("Bill batch input field not found or not visible")
                    current_date = datetime.datetime.now()
                    month_abbr = current_date.strftime("%b").upper()
                    year_short = current_date.strftime("%y")
                    sheet_name = getattr(self, 'current_sheet', 'Unknown')
                    self.append_log(f"🏷️ Determining bill name for sheet '{sheet_name}':")
                    self.append_log(f"   - Current WO Description: '{self.wo_description}'")
                    self.append_log(f"   - Excel Original Description: '{getattr(self, 'excel_original_description', 'N/A')}'")
                    description_for_bill = self.wo_description
                    if not description_for_bill or description_for_bill.strip() == "" or "To be stored" in description_for_bill or "سيتم التخزين" in description_for_bill:
                        if hasattr(self, 'excel_original_description') and self.excel_original_description:
                            description_for_bill = self.excel_original_description
                        else:
                            description_for_bill = f"WO-{self.wo_number}"
                    self.append_log(f"✅ Final description for bill name: '{description_for_bill}'")
                    bill_name = f"{description_for_bill}-{self.wo_number}-M{month_abbr}{year_short}"
                    bill_desc_input = page.locator('input[id="m9a223887-tb2"]').first
                    if bill_desc_input and bill_desc_input.is_visible():
                        bill_desc_input.fill(bill_name); self.append_log(f"Entered bill description: {bill_name}")
                        self.append_log("Looking for OK button...")
                        try:
                            ok_button = page.locator('button[id="m4e45caf4-pb"]').first
                            if ok_button.is_visible():
                                ok_button.click(); self.append_log("Successfully clicked OK button")
                            else:
                                for selector in ['button:has-text("OK")','button.text.pb.default:has-text("OK")','#m4e45caf4-pb']:
                                    try:
                                        alt_button = page.locator(selector).first
                                        if alt_button.is_visible():
                                            alt_button.click(); self.append_log(f"Clicked OK button using alternate selector: {selector}"); break
                                    except Exception as e:
                                        self.logger.warning(f"Failed to click OK with selector {selector}: {str(e)}")
                                else:
                                    raise Exception("Could not find or click OK button with any method")
                            page.wait_for_load_state("networkidle", timeout=30000); time.sleep(1)
                            self.append_log("Customer bill creation completed")
                        except Exception as e:
                            self.logger.error(f"Error clicking OK button: {str(e)}"); raise
                    else:
                        raise Exception("Bill description input field not found or not visible")
                except Exception as e:
                    self.append_log(f"Error processing bill batch form: {str(e)}", "error"); raise
            except Exception as e:
                self.append_log(f"Error in create_customer_bill: {str(e)}", "error"); raise
        except Exception:
            self.logger.exception("create_customer_bill failed")
            raise

    def handle_customer_bill(self, page):
        try:
            self.append_log("💰 Creating customer bill...")
            self.create_customer_bill(page)
            if self.stop_event.is_set():
                return
            self.save_bill_info(page)
            self.append_log(f"Customer bill created successfully with batch: {self.bill_batch}")
        except Exception:
            self.logger.exception("handle_customer_bill failed")
            raise

    def automate_maximo_tasks(self, page):
        """Process ONE sheet completely - full parity with Tkinter flow."""
        try:
            sheet_name = getattr(self, 'current_sheet', 'Unknown')
            self.append_log(f"🎯 Processing sheet '{sheet_name}' as a FRESH START")
            try:
                if not page.locator('#titlebar-tb_gotoButton').is_visible():
                    self.logger.warning("Login check failed, but continuing...")
            except Exception:
                pass
            self.append_log(f"🔍 Searching for Work Order: {self.wo_number}")
            self.search_work_order(page, self.wo_number)
            if self.stop_event.is_set():
                return
            try:
                page.wait_for_load_state("networkidle", timeout=30000)
                time.sleep(2)
                self.append_log(f"📝 Getting WO description for sheet '{sheet_name}' WO: {self.wo_number}")
                is_first_sheet = (sheet_name == 'Sheet1')
                if is_first_sheet:
                    page.wait_for_load_state("networkidle", timeout=10000); time.sleep(2)
                    self.append_log("📋 First sheet - using standard wait times")
                else:
                    page.wait_for_load_state("networkidle", timeout=15000); time.sleep(4)
                    self.append_log(f"📋 Subsequent sheet ({sheet_name}) - using extended wait times")
                    try:
                        page.wait_for_selector('.loading, .wait_modal, [id*="loading"]', state='hidden', timeout=5000)
                    except:
                        pass
                try:
                    description_selectors = [
                        '#mad3161b5-tb2',
                        'input[id="mad3161b5-tb2"]',
                        'input[maxlength="100"]',
                        'input[aria-labelledby*="description"]'
                    ]
                    page_description = None
                    max_retries = 3 if not is_first_sheet else 1
                    for retry in range(max_retries):
                        self.append_log(f"🔍 Description search attempt {retry + 1}/{max_retries}")
                        for selector in description_selectors:
                            try:
                                desc_element = page.locator(selector).first
                                if desc_element and desc_element.is_visible():
                                    desc_element.wait_for(state='visible', timeout=3000)
                                    time.sleep(0.5)
                                    page_description = desc_element.input_value()
                                    if page_description and page_description.strip():
                                        self.append_log(f"✅ Found description using selector: {selector} (attempt {retry + 1})")
                                        break
                                    else:
                                        self.append_log(f"⚠️ Element found but value is empty with selector: {selector}")
                            except Exception as sel_e:
                                self.append_log(f"⚠️ Selector {selector} failed: {str(sel_e)}")
                                continue
                        if page_description and page_description.strip():
                            break
                        elif retry < max_retries - 1:
                            self.append_log("🔄 Description not found, waiting 2 seconds before retry..."); time.sleep(2)
                    if page_description and page_description.strip():
                        self.wo_description = page_description.strip()
                        self.append_log(f"✅ Got ACTUAL description from Maximo: '{self.wo_description}'")
                        self.update_excel_description(self.wo_description)
                    else:
                        self.append_log("🔍 Primary selectors failed, trying comprehensive search...")
                        found_description = False
                        try:
                            all_inputs = page.locator('input[type="text"]').all()
                            self.append_log(f"🔍 Scanning {len(all_inputs)} text input fields...")
                            for idx, input_el in enumerate(all_inputs):
                                try:
                                    if input_el.is_visible():
                                        value = input_el.input_value()
                                        if value and len(value.strip()) > 5 and not value.strip().isdigit():
                                            if not value.startswith('WO') and 'WO' not in value and not value.startswith('DRAFT'):
                                                if not any(char.isdigit() for char in value[:3]):
                                                    self.wo_description = value.strip()
                                                    self.append_log(f"✅ Found description from input field {idx + 1}: '{self.wo_description}'")
                                                    self.update_excel_description(self.wo_description)
                                                    found_description = True
                                                    break
                                except Exception:
                                    continue
                            if not found_description:
                                self.append_log("🔍 Trying JavaScript evaluation for description...")
                                js_description = page.evaluate("""
                                    () => {
                                        const selectors = [
                                            'input[id*="description"]',
                                            'input[id*="desc"]',
                                            'input[maxlength="100"]',
                                            'textarea[id*="description"]'
                                        ];
                                        for (let selector of selectors) {
                                            const elements = document.querySelectorAll(selector);
                                            for (let el of elements) {
                                                if (el.value && el.value.trim().length > 5) { return el.value.trim(); }
                                            }
                                        }
                                        return null;
                                    }
                                """)
                                if js_description and js_description.strip():
                                    self.wo_description = js_description.strip()
                                    self.append_log(f"✅ Found description via JavaScript: '{self.wo_description}'")
                                    self.update_excel_description(self.wo_description)
                                    found_description = True
                            if not found_description:
                                self.append_log(f"⚠️ Could not find description field after comprehensive search")
                                self.append_log(f"📝 Keeping Excel description: '{self.wo_description}'")
                        except Exception as search_e:
                            self.append_log(f"⚠️ Error in comprehensive description search: {str(search_e)}", "warning")
                            self.append_log(f"📝 Will use Excel description for bill naming: '{self.wo_description}'")
                except Exception as e:
                    self.append_log(f"⚠️ Error getting description from page: {str(e)}", "warning")
                    self.append_log(f"📝 Will use Excel description for bill naming: '{self.wo_description}'")
                if not self.wo_description or self.wo_description.strip() == "":
                    if hasattr(self, 'excel_original_description') and self.excel_original_description:
                        self.wo_description = self.excel_original_description
                        self.append_log(f"📝 Using Excel original description: '{self.wo_description}'")
                    else:
                        self.wo_description = f"WO-{self.wo_number}"
                        self.append_log(f"⚠️ Using ultimate fallback description: '{self.wo_description}'")
                self.add_services_to_wo(page)
                if self.stop_event.is_set():
                    return
            except Exception as e:
                self.append_log(f"Error in work order navigation: {str(e)}", "error")
                raise
            sheet_name = getattr(self, 'current_sheet', 'Unknown')
            self.append_log(f"💰 Creating customer bill for WO: {self.wo_number} - {self.wo_description}")
            self.handle_customer_bill(page)
            if not self.bill_batch:
                self.append_log("⚠️ Warning: Bill batch number not generated!", "warning")
            self.append_log(f"✅ Completed automation for sheet '{sheet_name}' - WO: {self.wo_number}, Bill: {getattr(self, 'bill_batch', 'N/A')}")
        except Exception:
            self.logger.exception("automate_maximo_tasks failed")
            raise

# -----------------------
# Excel helpers
# -----------------------
def create_working_copy(uploaded_file) -> Optional[str]:
    try:
        # Persist uploaded file to disk with timestamp
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        original_name = Path(uploaded_file.name)
        base_dir = Path.cwd() / "output"
        base_dir.mkdir(parents=True, exist_ok=True)
        dest = base_dir / f"{original_name.stem}_working_{ts}{original_name.suffix}"
        with open(dest, "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.session_state["excel_path"] = str(dest)  # Treat as "selected" original
        st.session_state["working_excel_path"] = str(dest)  # Working copy same path for Streamlit
        append_log(f"📂 Created working copy at: {dest}")
        return str(dest)
    except Exception as e:
        update_status(f"Failed to create working copy: {e}", "error")
        return None

def load_excel_cache(path: str):
    try:
        st.session_state["excel_cache"] = pd.ExcelFile(path)
        st.session_state["total_sheets"] = len(st.session_state["excel_cache"].sheet_names)
        append_log("📄 Excel file loaded successfully")
        append_log(f"📄 Total sheets found: {st.session_state['total_sheets']}")
        append_log(f"📋 Sheet names: {', '.join(st.session_state['excel_cache'].sheet_names)}")
        return True
    except Exception as e:
        update_status(f"Error reading Excel: {e}", "error")
        return False

# -----------------------
# Playwright helpers (optional)
# -----------------------
def setup_browser(show_browser: bool):
    # Try lazy import to avoid false negatives from import at module load time
    try:
        from playwright.sync_api import sync_playwright as _sync_playwright
    except Exception as e:
        append_log(f"Playwright not available: {e}", "warning")
        return None, None, None, None
    try:
        playwright = _sync_playwright().start()
        # Determine if environment can show a browser window (Streamlit Cloud is headless)
        server_headless = os.environ.get("STREAMLIT_SERVER_HEADLESS", "true").lower() == "true"
        no_display = (sys.platform != "win32" and not os.environ.get("DISPLAY"))
        is_headless_env = server_headless or no_display
        headed = bool(show_browser) and not is_headless_env
        logging.getLogger("MaximoAutomation").info(f"Launching Chromium (headed={headed})")
        try:
            browser = playwright.chromium.launch(headless=not headed)
        except Exception as launch_err:
            # Attempt one-time install of Chromium if missing, then retry once
            msg = str(launch_err)
            logging.getLogger("MaximoAutomation").warning(f"Chromium launch failed, attempting install: {msg}")
            try:
                subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=False)
            except Exception:
                pass
            # retry
            browser = playwright.chromium.launch(headless=not headed)
        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            ignore_https_errors=True,
            locale='en-US',
            timezone_id='Asia/Riyadh'
        )
        page = context.new_page()
        append_log("🌐 Browser started")
        return playwright, browser, context, page
    except Exception as e:
        update_status(f"Failed to start browser: {e}", "error")
        return None, None, None, None

def close_browser(playwright, browser, context, page):
    try:
        if page: 
            try: page.close()
            except: pass
        if context:
            try: context.close()
            except: pass
        if browser:
            try: browser.close()
            except: pass
        if playwright:
            try: playwright.stop()
            except: pass
        append_log("Browser and playwright resources released")
    except Exception:
        pass

# -----------------------
# Validation
# -----------------------
def validate_inputs():
    # Always require an Excel working file
    if not st.session_state["working_excel_path"]:
        return False, "Please upload/select an Excel file"
    # Always require URL/credentials (no dry-run mode)
    if not st.session_state["url"]:
        return False, "Please enter the URL"
    if not st.session_state["url"].startswith(("http://", "https://")):
        return False, "Invalid URL format. Must start with http:// or https://"
    if not st.session_state["username"]:
        return False, "Please enter your username"
    if not st.session_state["password"]:
        return False, "Please enter your password"
    return True, ""

# -----------------------
# Main automation (threaded)
# -----------------------
def run_automation(params=None):
    playwright = browser = context = page = None
    try:
        # Controls already disabled in on_start; keep UI feedback lightweight here
        set_progress(0.0)
        update_status("Initializing automation...")

        # Pull required params captured on the UI thread to avoid session_state access issues in background threads
        if not isinstance(params, dict):
            update_status("Internal error: missing run parameters", "error")
            enable_controls()
            return
        excel_path = params.get("excel_path")
        url = params.get("url")
        username = params.get("username")
        password = params.get("password")
        selected_steps = params.get("selected_steps", ["add_services","route_to_comp","create_bill","put_prices"]) 
        start_at = params.get("start_at", "add_services")
        stop_event = params.get("stop_event")
        if not excel_path:
            update_status("Please upload/select an Excel file", "warning")
            enable_controls()
            return

        # Log Excel info (avoid st.session_state in threads)
        append_log(f"📂 Opening Excel file for processing: {Path(excel_path).name}")

        # Create automation adapter and bind optional cache/state
        automation = StreamlitAutomation(
            url=url,
            username=username,
            password=password,
            show_browser=True,
            remember=False,
            stop_event=stop_event,
            paused=False,
            excel_path=excel_path,
            working_excel_path=excel_path,
            excel_cache=None,
        )

        # Early stop check
        if stop_event and stop_event.is_set():
            update_status("Stopped before launching browser", "warning")
            enable_controls()
            return

        # Try to start browser (headed) regardless of module import at load time
        append_log(f"🔗 Opening login page: {url}")
        playwright, browser, context, page = setup_browser(True)
        automation.playwright, automation.browser, automation.context, automation.page = playwright, browser, context, page
        if page:
            try:
                page.goto(url, wait_until="networkidle")
            except Exception:
                pass

        # Perform login
        if page:
            automation.perform_login(page)

        # Process sheets with preserved flow
        try:
            xls = pd.ExcelFile(excel_path)
            sheet_names = xls.sheet_names
        except Exception as e:
            update_status(f"Failed reading Excel: {e}", "error")
            enable_controls()
            return
        total = len(sheet_names)
        append_log(f"🚀 Starting processing of {total} sheets")
        for idx, sheet in enumerate(sheet_names, 1):
            if stop_event and stop_event.is_set():
                update_status("Stopped between sheets", "warning")
                break

            automation.current_sheet = sheet
            try:
                st.session_state["current_sheet"] = sheet
            except Exception:
                pass
            update_status(f"Processing sheet {idx}/{total}: {sheet}")

            # Parse Excel data for this sheet (reuse original expectations)
            try:
                # Minimal extraction: header B2-B6 and services table like original
                # For parity, read the sheet via pandas and rebuild services_data similar to original.
                df = pd.read_excel(excel_path, sheet_name=sheet, header=None)
                # Header B2..B6 mapping
                try:
                    automation.wo_number = str(df.iloc[1, 1]).strip() if pd.notna(df.iloc[1, 1]) else None
                    automation.wo_description = str(df.iloc[2, 1]).strip() if pd.notna(df.iloc[2, 1]) else None
                    automation.bill_batch = str(df.iloc[3, 1]).strip() if pd.notna(df.iloc[3, 1]) else None
                    automation.bill_status = str(df.iloc[4, 1]).strip() if pd.notna(df.iloc[4, 1]) else None
                    automation.reviewed_by = str(df.iloc[5, 1]).strip() if pd.notna(df.iloc[5, 1]) else None
                except Exception:
                    pass

                # Locate services header row by finding a row containing 'Service Item'
                header_row_idx = None
                for r in range(len(df)):
                    row_vals = df.iloc[r].astype(str).str.strip().str.lower().tolist()
                    if any(val == 'service item' for val in row_vals):
                        header_row_idx = r
                        break
                services_df = None
                if header_row_idx is not None:
                    services_df = df.iloc[header_row_idx:].copy()
                    services_df.columns = services_df.iloc[0]
                    services_df = services_df.iloc[1:].copy()
                    services_df.columns = services_df.columns.astype(str).str.strip()
                    # Normalize column names used by automation (broader mapping)
                    rename_map = {}
                    for col in services_df.columns:
                        cl = str(col).strip().lower()
                        if cl in ["service item", "serviceitem", "serial", "service_item", "item", "service code", "code"]:
                            rename_map[col] = "Service Item"
                        elif cl in ["total price", "total_price", "total", "amount", "line total"]:
                            rename_map[col] = "Total Price"
                        elif cl in ["description", "desc", "service description"]:
                            rename_map[col] = "Description"
                        elif cl in ["quantity", "qty", "qyt", "qty.", "qnty"]:
                            rename_map[col] = "Quantity"
                        elif cl in ["unit price", "unit_price", "price", "rate"]:
                            rename_map[col] = "Unit Price"
                    if rename_map:
                        services_df = services_df.rename(columns=rename_map)
                    # Keep canonical columns if present
                    keep_cols = [c for c in ["Service Item", "Description", "Quantity", "Unit Price", "Total Price"] if c in services_df.columns]
                    if keep_cols:
                        services_df = services_df[keep_cols]
                automation.services_data = services_df if services_df is not None else pd.DataFrame()
            except Exception as e:
                append_log(f"Sheet '{sheet}' parse error: {e}", "warning")
                automation.services_data = pd.DataFrame()

            # Validate minimal requirements
            if not automation.wo_number:
                append_log(f"❌ Missing WO number for sheet '{sheet}', skipping.", "warning")
                set_progress(idx/total)
                continue

            # Update UI-hint fields
            try:
                st.session_state["last_wo_number"] = automation.wo_number or ""
            except Exception:
                pass

            # Step pipeline: run selected steps starting from chosen start point
            canonical = ["add_services","route_to_comp","create_bill","put_prices"]
            try:
                start_idx = canonical.index(start_at)
            except ValueError:
                start_idx = 0
            ordered = [s for s in canonical[start_idx:] if s in selected_steps]

            if page:
                # Ensure we're on the WO first
                automation.search_work_order(page, automation.wo_number)
                for step in ordered:
                    if stop_event and stop_event.is_set():
                        update_status("Stop requested; exiting current sheet.", "warning")
                        break
                    if step == "add_services":
                        automation.add_services_to_wo(page)
                    elif step == "route_to_comp":
                        automation.route_wo_to_comp(page)
                    elif step == "create_bill":
                        automation.create_customer_bill(page)
                        # try to capture and persist bill info if available
                        try:
                            automation.save_bill_info(page)
                        except Exception:
                            pass
                        try:
                            st.session_state["last_bill_batch"] = automation.bill_batch or st.session_state.get("last_bill_batch", "")
                        except Exception:
                            pass
                    elif step == "put_prices":
                        automation.process_customer_bill(page)
                        try:
                            automation.save_bill_changes(page)
                        except Exception:
                            pass
                # Always reflect back to Excel after running selected steps
                automation.update_excel_file()
            else:
                # No browser available; still sync Excel info
                automation.update_excel_file()

            append_log(f"✅ Sheet '{sheet}' processed")
            set_progress(idx/total)

        update_status("Automation complete", "info")
        append_log("🎉 All done.")
    except Exception as e:
        update_status(f"Automation error: {e}", "error")
        logging.getLogger("MaximoAutomation").exception("Automation error")
    finally:
        try:
            close_browser(playwright, browser, context, page)
        except Exception:
            pass
        enable_controls()

# -----------------------
# Event handlers
# -----------------------
def on_start():
    ok, msg = validate_inputs()
    if not ok:
        update_status(msg, "warning")
        return
    # Immediately reflect running state and status for user feedback
    disable_controls()
    update_status("Starting automation...", "info")
    st.session_state["stop_event"] = threading.Event()
    st.session_state["paused"] = False
    st.session_state["logs"] = []
    append_log("🚀 Starting automation process...")
    append_log(f"🌍 URL: {st.session_state['url']}")
    append_log(f"👤 Username: {st.session_state['username']}")
    if st.session_state["working_excel_path"]:
        append_log(f"📂 Excel file: {Path(st.session_state['working_excel_path']).name}")

    # Capture params to avoid accessing st.session_state inside worker thread
    params = {
        "excel_path": st.session_state.get("working_excel_path"),
        "url": st.session_state.get("url"),
        "username": st.session_state.get("username"),
        "password": st.session_state.get("password"),
        "selected_steps": st.session_state.get("selected_steps", [
            "add_services","route_to_comp","create_bill","put_prices"
        ]),
        "start_at": st.session_state.get("start_at", "add_services"),
        "stop_event": st.session_state.get("stop_event"),
    }

    t = threading.Thread(target=run_automation, args=(params,), name="AutomationThread", daemon=True)
    st.session_state["automation_thread"] = t
    t.start()

def on_pause():
    st.session_state["paused"] = not st.session_state["paused"]
    if st.session_state["paused"]:
        append_log("Automation paused")
    else:
        append_log("Automation resumed")

def on_stop():
    if st.session_state.get("stop_event"):
        st.session_state["stop_event"].set()
        update_status("Stop requested; will stop safely at next checkpoint.", "warning")
        append_log("Stop event set by user")

def on_save_logs():
    buffer = io.StringIO()
    for line in st.session_state.get("logs", []):
        buffer.write(line + "\n")
    st.download_button(
        label="Download logs as .txt",
        data=buffer.getvalue().encode("utf-8"),
        file_name="maximo_logs.txt",
        mime="text/plain",
        key="download_logs_btn"
    )

# -----------------------
# Sidebar (env, preferences)
# -----------------------
with st.sidebar:
    st.markdown("### Environment")
    # Use a single radio to avoid widget-state mutation conflicts
    env_choice = st.radio(
        "Select Environment",
        options=("DEV", "PROD"),
        index=0 if st.session_state.get("dev_env", True) else 1,
        horizontal=True,
        key="env_choice",
    )
    # Reflect radio choice into our internal flags (no conflicting widget keys)
    st.session_state["dev_env"] = env_choice == "DEV"
    st.session_state["prod_env"] = env_choice == "PROD"
    st.session_state["url"] = DEV_URL if st.session_state["dev_env"] else PROD_URL

    st.markdown("---")
    st.markdown("### Preferences")
    st.session_state["show_password"] = st.checkbox("Show Password", value=st.session_state["show_password"])
    # Always show browser; remove toggles that hide it or skip it
    st.session_state["show_browser"] = True
    st.session_state["dry_run"] = False
    st.session_state["remember"] = st.checkbox("Remember credentials", value=st.session_state["remember"])
    if st.session_state["remember"]:
        # Save immediately when toggled on
        save_credentials(st.session_state["username"], st.session_state["password"], st.session_state["url"], True)

# -----------------------
# Main UI (cards to mimic Tk layout)
# -----------------------
st.markdown(f"## {APP_TITLE}")

# Login card
with st.container():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">🔑 Login Information</div>', unsafe_allow_html=True)
    st.session_state["url"] = st.text_input("URL:", value=st.session_state["url"], help="Maximo login URL")
    st.session_state["username"] = st.text_input("Username:", value=st.session_state["username"])
    st.session_state["password"] = st.text_input(
        "Password:", value=st.session_state["password"], type="default" if st.session_state["show_password"] else "password"
    )
    st.markdown('</div>', unsafe_allow_html=True)

# File card
with st.container():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📂 Excel File</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Select Excel File (.xlsx, .xlsm)", type=["xlsx", "xlsm"])
    if uploaded is not None:
        path = create_working_copy(uploaded)
        if path:
            # Load cache once selected
            load_excel_cache(path)
    if st.session_state["working_excel_path"]:
        st.caption(f"Working file: {st.session_state['working_excel_path']}")
    st.markdown('</div>', unsafe_allow_html=True)

# Controls
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">🎛️ Controls</div>', unsafe_allow_html=True)

# Step selection controls
step_labels = [
    ("Add Services", "add_services"),
    ("Route to COMP", "route_to_comp"),
    ("Create Bill", "create_bill"),
    ("Put Bill Prices", "put_prices"),
]
label_to_code = {lbl: code for (lbl, code) in step_labels}
code_to_label = {code: lbl for (lbl, code) in step_labels}

selected_labels = st.multiselect(
    "Steps to run (in order):",
    options=[lbl for (lbl, _) in step_labels],
    default=[code_to_label[c] for c in st.session_state.get("selected_steps", []) if c in code_to_label],
)
selected_codes = [label_to_code[lbl] for lbl in selected_labels]
st.session_state["selected_steps"] = selected_codes if selected_codes else st.session_state.get("selected_steps", [])

start_label = st.selectbox(
    "Start at:",
    options=[lbl for (lbl, _) in step_labels],
    index=[lbl for (lbl, _) in step_labels].index(code_to_label.get(st.session_state.get("start_at", "add_services"), "Add Services")),
)
st.session_state["start_at"] = label_to_code[start_label]

ctrl_col1, ctrl_col2, ctrl_col3, ctrl_col4 = st.columns([1,1,1,1])
with ctrl_col1:
    st.button("▶️ Start", on_click=on_start, disabled=st.session_state["is_running"]) 
with ctrl_col2:
    st.button("⏸️ Pause", on_click=on_pause, disabled=not st.session_state["is_running"]) 
with ctrl_col3:
    st.button("⏹️ Stop", on_click=on_stop, disabled=not st.session_state["is_running"]) 
with ctrl_col4:
    if st.button("💾 Save Logs"):
        on_save_logs()
st.markdown('</div>', unsafe_allow_html=True)

# Progress + Status
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">📈 Progress</div>', unsafe_allow_html=True)
st.progress(st.session_state["progress"]) 
st.markdown(f'<span class="progress-text">{st.session_state["progress_text"]}</span>', unsafe_allow_html=True)
# Inline current sheet and IDs
if st.session_state.get("current_sheet"):
    st.caption(f"Sheet: {st.session_state['current_sheet']}")
wo_hint = st.session_state.get("last_wo_number") or ""
batch_hint = st.session_state.get("last_bill_batch") or ""
if wo_hint or batch_hint:
    st.caption(f"WO: {wo_hint}  |  Bill Batch: {batch_hint}")
if st.session_state["status"]:
    st.info(st.session_state["status"]) 
st.markdown('</div>', unsafe_allow_html=True)

# Logs
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">🧾 Logs</div>', unsafe_allow_html=True)
lines = st.session_state.get("logs", [])
colored = []
for ln in lines:
    lvl = "info"
    u = ln.upper()
    if "- ERROR -" in u or u.startswith("ERROR"):
        lvl = "error"
    elif "- WARNING -" in u or u.startswith("WARNING"):
        lvl = "warning"
    colored.append(f'<span class="log-line log-{lvl}">{ln}</span>')
html_logs = "".join(colored)
st.markdown(f'<div class="log-container">{html_logs}</div>', unsafe_allow_html=True)
st.markdown('<div class="small-muted">Logs are also written to your user folder under maximo_logs.</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# Notes:
# - This app mirrors the layout and flow of the Tkinter app as closely as possible using Streamlit paradigms.
# - Long-running automation executes in a background thread; UI remains responsive.
# - Replace the simulated work inside run_automation() with your exact routines from backup_Stream_30.py as needed.
# - For Playwright usage, ensure it is installed and browsers are set up:
#     pip install playwright
#     python -m playwright install chromium

